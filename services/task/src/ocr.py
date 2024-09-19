import cv2
import pandas as pd
from paddleocr import PaddleOCR, PPStructure
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from pprint import pprint
import tempfile

from src.models.ocr_model import OCRResult, OCRResponse


def perform_paddle_ocr(ocr: PaddleOCR, image_path: Path) -> OCRResponse:
    raw_results = ocr.ocr(str(image_path))
    ocr_results = [
        OCRResult(
            bounding_box=result[0],
            text=result[1][0],
            confidence=result[1][1]
        )
        for result in raw_results[0]
    ]
    return OCRResponse(results=ocr_results)

import tempfile

def process_image_and_create_excel(table_engine: PPStructure, image_path: Path) -> Path:
    # Create a temporary file for the Excel output
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        output_path = Path(temp_file.name)

    # Create and save an Excel workbook to store the results
    Workbook().save(output_path)
    book = load_workbook(output_path)
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    writer.book = book

    # Process the image
    print(f'Processing image')
    img = cv2.imread(str(image_path))
    result = table_engine(img)
    pprint(result)

    # Create an image object for openpyxl
    xlimg = XLImage(str(image_path))

    i = 1
    for line in result:
        # Remove the 'img' key from the result
        line.pop('img')
        # Check if the line is a table
        if line.get("type") == "table":
            # Extract HTML table and convert to DataFrame
            html_table = line.get("res").get("html")
            html_data = pd.read_html(html_table)
            df = pd.DataFrame(html_data[0])

            # Write DataFrame to Excel and add the image to the sheet
            df.to_excel(writer, sheet_name=f"table {i}", index=1)
            book[f"table {i}"].add_image(xlimg, 'A100')
            i += 1

    # Save the Excel workbook
    writer.save()

    return output_path